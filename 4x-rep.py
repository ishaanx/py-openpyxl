#Import required libraries
import os
import xlsxwriter
import glob
import csv
import shutil
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from time import sleep
import string
from alive_progress import alive_bar
import sys
import subprocess
import time
import openpyxl
import openpyxl.styles
import openpyxl.utils
import openpyxl.workbook
from myVariables import *



def cc_sales():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing CC Sales Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_cc_sales
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)
    #print(wd)
    #print(wd_name)
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"

    # print(prev_mon)
    fd_name = "./Export/" + "Sales" + prev_mon + ".xlsx"
    #print(fd_name)
    # main prog
    with alive_bar(
        total=100,
        manual="True",
        title="CC Sales",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        # convert csv to xlsx using pandas lib
        colnames = ["Request ID",
        "Confirmation Number",
        "Status",
        "Payment Date",
        "Payment Amount",
        "UserID",
        "UserName",
        "Guest Name",
        "Source Code",
        "Source Name",
        "CP Code",
        "CP Name",
        "LDB code",
        "LDB Name",
        "Booking Date",
        "Booking Time GMT",
        "Checked In Date",
        "External Confirmation No",
        "Room Rent",
        "Checked-in Time",
        "Remark"]
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep=",",  header=None,  names=colnames, encoding='utf-8',low_memory=False)
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep=",",  header=None,  names=colnames, encoding='cp1252',low_memory=False)
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)
        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active

        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def cc_refunds():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing CC Refunds Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_cc_refunds
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)
    #print(wd)
    #print(wd_name)
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"

    # print(prev_mon)
    fd_name = "./Export/" + "Refunds" + prev_mon + ".xlsx"
    #print(fd_name)
    # main prog
    with alive_bar(
        total=100,
        manual="True",
        title="CC Refunds",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        # convert csv to xlsx using pandas lib
        colnames = ["Request ID",
        "Confirmation Number",
        "Status",
        "Payment Date",
        "Amount",
        "UserID",
        "UserName",
        "Guest Name",
        "Source Code",
        "Source Name",
        "CP Code",
        "CP Name",
        "LDB code",
        "LDB Name",
        "Booking Date",
        "Booking Time GMT",
        "Checked In Date",
        "External Confirmation No",
        "Room Rent",
        "Checked-in Time",
        "Remarks"]
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep=",",  header=None,  names=colnames, encoding='utf-8',low_memory=False)
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep=",",  header=None,  names=colnames, encoding='cp1252',low_memory=False)
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)
        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active

        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def payments():
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_payments
    fs_ext = ".tsv"
    fs_file_name = fs_path + "/" + fs_name + fs_ext

    filehandler_path = fs_file_name
    print(filehandler_path)
    output_path = "./temp"

    def cleanup():
        with alive_bar(
            title="Cleaning up old files",
            theme="smooth",
            bar="blocks",
            spinner="classic",
        ) as bar:
            if os.path.exists(output_path):
                shutil.rmtree(output_path)
                bar()

    def split(
        filehandler=filehandler_path,
        delimiter="\t",
        row_limit=300000,
        output_name_template="Payments_%s.tsv",
        keep_headers=True,
    ):
        """
        Splits a CSV file into multiple pieces.
        Arguments:
            `row_limit`: The number of rows you want in each output file. 10,000 by default.
            `output_name_template`: A %s-style template for the numbered output files.
            `output_path`: Where to stick the output files.
            `keep_headers`: Whether or not to print the headers in each output file.
        """
        print("\n")
        with alive_bar(
            total=100,
            manual=True,
            title="Splitting csv",
            theme="smooth",
            bar="blocks",
            spinner="classic",
        ) as bar1:
            # print ('Creating Export paths')
            if not os.path.exists(output_path):
                os.makedirs(output_path)
            bar1(0.10)
            if not os.path.exists("Export"):
                os.makedirs("Export")
            bar1(0.20)
            try:
                reader = csv.reader(open(filehandler, "r", encoding='utf-8'), delimiter=delimiter)
            except ValueError:
                reader = csv.reader(open(filehandler, "r", encoding='cp1252'), delimiter=delimiter)
            else:
                reader = csv.reader(open(filehandler, "r", encoding='latin1'), delimiter=delimiter)                
            current_piece = 1
            current_out_path = os.path.join(
                output_path, output_name_template % current_piece
            )
            bar1(0.30)
            current_out_writer = csv.writer(
                open(current_out_path, "w"), delimiter=delimiter
            )
            current_limit = row_limit
            if keep_headers:
                headers = next(reader)
                current_out_writer.writerow(headers)
            bar1(0.40)
            for i, row in enumerate(reader):
                if i + 1 > current_limit:
                    current_piece += 1
                    current_limit = row_limit * current_piece
                    current_out_path = os.path.join(
                        output_path, output_name_template % current_piece
                    )
                    current_out_writer = csv.writer(
                        open(current_out_path, "w"), delimiter=delimiter
                    )
                    bar1(0.50)
                    bar1(0.60)
                    if keep_headers:
                        current_out_writer.writerow(headers)
                    bar1(0.70)
                current_out_writer.writerow(row)
            bar1(0.80)
            bar1(0.90)
            bar1(1)
            return filehandler

    # join multiple csv to a single workbook with diff worksheets
    def split_join():
        from pathlib import Path

        # Returns the same day of last month if possible otherwise end of month
        # (eg: March 31st->29th Feb an July 31st->June 30th)
        last_month = datetime.now() - relativedelta(months=1)
        # Create string of month name and year...
        text = format(last_month, "%B %Y")
        prev_mon = "[" + text + "]"
        fname = "./Export/" + "Payments" + prev_mon + v_api + ".xlsx"
        writer = pd.ExcelWriter(fname, engine="xlsxwriter")
        folders = "temp"
        print("\n")
        with alive_bar(
            title="Combining csv", theme="smooth", bar="blocks", spinner="classic"
        ) as bar2:
            myPath = os.path.join(os.getcwd(), folders)
            for myPath in Path(myPath).rglob("*.tsv"):
                df = pd.read_csv(myPath, sep="\t", low_memory=False)
                bar2()
                df.to_excel(
                    writer, index=False, sheet_name=os.path.basename(myPath)[:31]
                )
            writer.save()
        print("\n")
        return fname

    def styl():
        import openpyxl
        from openpyxl.reader.excel import load_workbook
        from openpyxl.styles import Font, Alignment
        from openpyxl.utils import get_column_letter

        # Returns the same day of last month if possible otherwise end of month
        # (eg: March 31st->29th Feb an July 31st->June 30th)
        last_month = datetime.now() - relativedelta(months=1)
        # Create string of month name and year...
        text = format(last_month, "%B %Y")
        prev_mon = "[" + text + "]"
        fname = "./Export/" + "Payments" + prev_mon + v_api + ".xlsx"
        wb = load_workbook(fname)
        for ws in wb.worksheets:
            mr = ws.max_row
            mc = ws.max_column
            with alive_bar(
                title="Applying Font Styles",
                theme="smooth",
                bar="blocks",
                spinner="classic",
            ) as bar3:
                for cell in ws["mr:mc"]:
                    cell.font = Font(size=11)
                    bar3()
                for cell in ws["1:1"]:
                    cell.font = Font(size=12)
                    cell.style = "Accent1"
                    cell.alignment = Alignment(wrapText="True", horizontal="center")
                    # bar4()
                for col in range(1, 30):
                    ws.column_dimensions[(get_column_letter(col))].width = 15
                    ws.freeze_panes = "A2"
                    # bar5()
        print("\n")
        wb.save(fname)

    cleanup()
    split()
    split_join()
    styl()
    cleanup()


def chg_and_adj():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing Charges and Adjustments Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_chg_and_adj
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)
    #print(wd)
    #print(wd_name)
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"

    # print(prev_mon)
    fd_name = "./Export/" + "Charges and Adjustments" + prev_mon + v_api + ".xlsx"
    #print(fd_name)
    # main prog
    with alive_bar(
        total=100,
        manual="True",
        title="Charges and Adjustments",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        #### Replace double quotes with whitespace
        def spl_replace():
          with open(fs_path + "/" + fs_name + fs_ext, 'r', encoding="cp1252") as f:
            text = f.read()

          converted_text = text.replace('"', " ")

          with open(fs_path + "/" + fs_name + fs_ext, 'w', encoding="cp1252") as f:
            f.write(converted_text)

        spl_replace()
        # convert csv to xlsx using pandas lib
        colnames = ["Property Code","Confirmation No","Guest Name","Check in Date","Check in Time","Check out Date","Check out Time"," Room Number","Charge Date","Charge Created at Date","Charge Created at Time","Charge Name","Adjustment Date","Adjustment Created at Date","Adjustment Created at Time","Adjustment Amount","Charge Rate Code Old","Charge Rate Code New","Reason Code","Username","User","Reservation Status","Remarks"]
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",  header=None,  names=colnames, encoding='utf-8', skiprows=1 )
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",  header=None,  names=colnames, encoding='cp1252', skiprows=1 )
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)
        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active


        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def cct():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("\n")
    print("Processing CCT Info Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_cct
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)
        
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"

    # print(prev_mon)
    fd_name = "./Export/" + "CCT Information" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="CCT Information",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        # convert csv to xlsx using pandas lib
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='utf-8')
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='cp1252')
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)

        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        # Columns: property_code date  serial_no   product applicationVersion  manufacturer    mfg_serial_no
        ws["A1"] = "Property Code"
        ws["B1"] = "Date"
        ws["C1"] = "Serial No"
        ws["D1"] = "Product"
        ws["E1"] = "Application Version"
        ws["F1"] = "Manufacturer"
        ws["G1"] = "MFG Serial No"

        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def discp_rates():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing Discrepant Rates")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_discp_rates
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)

    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"

    # print(prev_mon)
    fd_name = "./Export/" + "Discrepant Rates" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="Discrepant Rates",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        # convert csv to xlsx using pandas lib
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='utf-8')
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='cp1252')
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)

        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def dnr1():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing DNR 1 Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_dnr1
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)

    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"

    fd_name = "./Export/" + "DNR 1" + prev_mon + v_api + ".xlsx"
    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="DNR 1",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        # convert csv to xlsx using pandas lib
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='utf-8')
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='cp1252')
        bar(0.10)
        #read_file.to_excel("" r"" + fd_name, index=None, header=True)
        read_file.to_excel(fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)

        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def dnr2():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing DNR 2 Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_dnr2
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)

    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"

    # print(prev_mon)
    fd_name = "./Export/" + "DNR 2" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="DNR 1",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        # convert csv to xlsx using pandas lib
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='utf-8')
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='cp1252')
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)

        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def dnr3():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing DNR 3 Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_dnr3
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)

    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"

    # print(prev_mon)
    fd_name = "./Export/" + "DNR 3" + prev_mon + v_api + ".xlsx"
    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="DNR 1",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        # convert csv to xlsx using pandas lib
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='utf-8')
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='cp1252')
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)

        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def grts_and_gst():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing GRATIS & GSTCERT Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_grts_and_gst
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)

    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"

    # print(prev_mon)
    fd_name = "./Export/" + "GRATIS and GSTCERT" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="GRATIS and GSTCERT 1",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        # convert csv to xlsx using pandas lib
        try:
            read_file = pd.read_csv(
                ""r"" + fs_path + "/" + fs_name + fs_ext,
                sep="\t",
                header=0,
                encoding='utf-8',
                names=[
                    "Property Code",
                    "Confirmation Number",
                    "Checkin Date",
                    "Checkout Date",
                    "Number of Nights",
                    "Payment Tpe",
                    "Guest Name",
                    "Number of Payments",
                ],
            )
        except ValueError:
            read_file = pd.read_csv(
                ""r"" + fs_path + "/" + fs_name + fs_ext,
                sep="\t",
                header=0,
                encoding='cp1252',
                names=[
                    "Property Code",
                    "Confirmation Number",
                    "Checkin Date",
                    "Checkout Date",
                    "Number of Nights",
                    "Payment Tpe",
                    "Guest Name",
                    "Number of Payments",
                ],
            )                        
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)

        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def gst_email():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing Guest Email Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_gst_email
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)

    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"
    # print(prev_mon)
    fd_name = "./Export/" + "Guest Email" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="Guest Email",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        # convert csv to xlsx using pandas lib
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='utf-8')
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='cp1252')
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)

        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def lldb():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing LLDB Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_lldb
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)
        
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"
    # print(prev_mon)
    fd_name = "./Export/" + "LLDB" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="LLDB",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        #### Replace double quotes with whitespace
        def spl_replace():
          with open(fs_path + "/" + fs_name + fs_ext, 'r', encoding="cp1252") as f:
            text = f.read()

          converted_text = text.replace('"', " ")

          with open(fs_path + "/" + fs_name + fs_ext, 'w', encoding="cp1252") as f:
            f.write(converted_text)

        spl_replace()
        # convert csv to xlsx using pandas lib
        read_file = pd.read_csv(
            "" r"" + fs_path + "/" + fs_name + fs_ext,
            sep="\t",
            header=0,
            encoding='cp1252',
            names=[
                "Property Code",
                "Property Name",
                "Property Type",
                "Property Brand",
                "State",
                "Company Name",
                "Company Code",
                "Confirmation Number",
                "Checkin Date",
                "Checkout Date",
                "Number of Nights",
                "Total Charges",
                "External Confirmation No",
                "Booking Source",
                "Rate Code",
            ],
        )
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)

        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def pay_and_ref():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing Payments & Refunds Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_pay_and_ref
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)


    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"
    # print(prev_mon)
    fd_name = "./Export/" + "Payments and Refunds" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="Payments and Refunds",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        #### Replace double quotes with whitespace
        def spl_replace():
          with open(fs_path + "/" + fs_name + fs_ext, 'r', encoding="cp1252") as f:
            text = f.read()

          converted_text = text.replace('"', " ")

          with open(fs_path + "/" + fs_name + fs_ext, 'w', encoding="cp1252") as f:
            f.write(converted_text)

        spl_replace()        
        # convert csv to xlsx using pandas lib
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='utf-8')
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='cp1252')
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)

        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def prop_over():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing Property Overview Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_prop_over
    fs_ext = ".tsv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)

    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"
    # print(prev_mon)
    fd_name = "./Export/" + "Property Overview" + prev_mon + v_api + ".xlsx"
    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="Property Overview",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting

        ## Report 1 -
        #### Replace double quotes with whitespace
        def spl_replace():
          with open(fs_path + "/" + fs_name + fs_ext, 'r', encoding="cp1252") as f:
            text = f.read()

          converted_text = text.replace('"', " ")

          with open(fs_path + "/" + fs_name + fs_ext, 'w', encoding="cp1252") as f:
            f.write(converted_text)

        spl_replace()        
        # convert csv to xlsx using pandas lib
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='utf-8')
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t",encoding='cp1252')
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)

        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def room_moves():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing Room Moves Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_room_moves
    fs_ext = ".tsv"
    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"
    # print(prev_mon)
    fd_name = "./Export/" + "Room Moves" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="Room Moves",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  # default setting
        ## Report 1 -
        # convert csv to xlsx using pandas lib

        #### Replace double quotes with whitespace
        def spl_replace():
          with open(fs_path + "/" + fs_name + fs_ext, 'r', encoding="cp1252") as f:
            text = f.read()

          converted_text = text.replace('"', " ")

          with open(fs_path + "/" + fs_name + fs_ext, 'w', encoding="cp1252") as f:
            f.write(converted_text)

        spl_replace()
        # convert csv to xlsx using pandas lib
        colnames2 = ["Property","Confirmation Number","Guest Name","Checked In Date","Checked In Time", "From Room","To Room", "Move Date","Move Time","Remarks", "User"]
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t", header=None,skiprows=1,  names=colnames2, encoding='utf-8')
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t", header=None,skiprows=1,  names=colnames2, encoding='cp1252')
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)
        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def os_users():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    import numpy as np

    print("Processing Organization Structure Users Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_os_users
    fs_ext = ".tsv"
    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"
    # print(prev_mon)
    fd_name = "./Export/" + "OS Users" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="OS Users",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  

        # convert csv to xlsx using pandas lib
        colnames2 = ["Organization Structure", "Parent", "User Template", "Enterprise Template", "First Name", "Last Name", "Username"]
        read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t", header=None,skiprows=1,  names=colnames2)
        read_file.fillna("NULL",inplace=True) ##Replaces NaN with "NULL" string
        # print(read_file)
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)
        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def os_properties():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing Organization Structure Properties Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_os_properties
    fs_ext = ".tsv"
    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"
    # print(prev_mon)
    fd_name = "./Export/" + "OS Properties" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="OS Properties",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  

        # convert csv to xlsx using pandas lib
        colnames2 = ["Organization Structure", "Parent", "User Template", "Enterprise Template", "Property Code"]
        read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t", header=None,skiprows=1,  names=colnames2)
        read_file.fillna("NULL",inplace=True) ##Replaces NaN with "NULL" string
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)
        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


def all_users():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter

    print("Processing All Users Report")
    ## DECLARE VARIABLES
    # Source vars
    fs_path = os.getcwd()
    fs_name = v_all_users
    fs_ext = ".tsv"
    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    # exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, "%B %Y")
    prev_mon = "[" + text + "]"
    # print(prev_mon)
    fd_name = "./Export/" + "All Users" + prev_mon + v_api + ".xlsx"

    # main prog
    with alive_bar(
        total=100,
        manual=True,
        title="All Users",
        theme="smooth",
        bar="blocks",
        spinner="classic",
    ) as bar:  

        # convert csv to xlsx using pandas lib
        colnames2 = ["Property Assigned", "First Name", "Last Name", "Username", "User Status", "Enterprise Template", "User Template"]
        try:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t", header=None,skiprows=1,  names=colnames2, encoding='utf-8',low_memory=False)
        except ValueError:
            read_file = pd.read_csv("" r"" + fs_path + "/" + fs_name + fs_ext, sep="\t", header=None,skiprows=1,  names=colnames2, encoding='cp1252',low_memory=False)
        read_file.fillna("NULL",inplace=True) ##Replaces NaN with "NULL" string
        bar(0.10)
        read_file.to_excel("" r"" + fd_name, index=None, header=True)
        bar(0.20)
        # read xlsx
        # assign the excel file to wb() variable
        wb = load_workbook(fd_name)
        bar(0.30)
        # assign the worksheet of the workbook to a ws() variable
        ws = wb.active
        bar(0.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws["mr:mc"]:
            cell.font = Font(size=11)
        bar(0.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = "Accent1"
            cell.alignment = Alignment(wrapText="True", horizontal="center")
        bar(0.80)
        # set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        # Save the excel file
        bar(1)
        wb.save(fd_name)

    print("Export Completed\n")


## Cleanup Function to remove export folder
def clean():
    if os.path.exists("Export"):
        shutil.rmtree("Export")

## If user input = all then execute all functions listed below
def all():
    print("Processing all reports")
    clean()
    payments()
    cct()
    chg_and_adj()
    discp_rates()
    dnr1()
    dnr2()
    dnr3()
    grts_and_gst()
    gst_email()
    lldb()
    pay_and_ref()
    prop_over()
    room_moves()
    os_users()
    os_properties()
    all_users()

## Dispatcher function used to select user choice
dispatcher = {
    "1": payments,
    "2": cct,
    "3": chg_and_adj,
    "4": discp_rates,
    "5": dnr1,
    "6": dnr2,
    "7": dnr3,
    "8": grts_and_gst,
    "9": gst_email,
    "10": lldb,
    "11": pay_and_ref,
    "12": prop_over,
    "13": room_moves,
    "14": os_users,
    "15": os_properties,
    "16": all_users,
    "all": all,
}
os.system('cls' if os.name == 'nt' else 'clear')
print("Created by ishan badgainya")
print(
    "Following choices are available:\n\
    1 - Payments Report\n\
    2 - CCT Information Report\n\
    3 - Charges and Adjustments Report\n\
    4 - Discrepant Rates Report\n\
    5 - DNR 1 Report\n\
    6 - DNR 2 Report\n\
    7 - DNR 3 Report\n\
    8 - GRTS and GSTCRT Report\n\
    9 - Guest Email Report\n\
    10 - LLDB Report\n\
    11 - Payments and Refunds Report\n\
    12 - Property Overview Report\n\
    13 - Room Moves Report\n\
    14 - Organization Structure Users Report\n\
    15 - Organization Structure Properties Report\n\
    16 - All Users Report\n\
    "
)
action = input("Enter a number [1 to 16] or 'all' to process all reports: - ")
dispatcher[action]()
