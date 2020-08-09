import os
import xlsxwriter
import glob
import csv
import shutil
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from time import sleep
import string
from alive_progress import alive_bar
import sys
import subprocess
import time
import openpyxl
import openpyxl.styles
import openpyxl.utils
import openpyxl.workbook

#Input file names without extension
#Example: 'Payments-Jun'
v_chg_and_adj = 'Charges_And_Adjustment_Jun'
v_payments = 'Payments-Jun'
v_cct = 'CCT-Information-Jun'
v_discp_rates = 'Discrepant_Rates_Jun'
v_dnr1 = 'DNR-1-Jun'
v_dnr2 = 'DNR-2-Jun'
v_dnr3 = 'DNR-3-Jun'
v_grts_and_gst = 'Gratis_And_GstCert_Report_Jun'
v_gst_email = 'Guest_Email_Jun'
v_lldb = 'LLDB_Jun'
v_pay_and_ref = 'Payments_And_Refunds_Jun'
v_prop_over = 'Property_Overview_Jun'
v_room_moves = 'Room_Moves_Jun'

def payments():

    #print('Running one function')
    filehandler_path ='Payments-Jun.csv'
    print(filehandler_path)
    output_path='./temp'

    def cleanup():
        print('Cleaning Up temp files')
        if os.path.exists(output_path):
            shutil.rmtree(output_path)


    def split(filehandler=filehandler_path, delimiter=',', row_limit=300000, 
        output_name_template='Payments_%s.csv',  keep_headers=True):
        """
        Splits a CSV file into multiple pieces.
        Arguments:
            `row_limit`: The number of rows you want in each output file. 10,000 by default.
            `output_name_template`: A %s-style template for the numbered output files.
            `output_path`: Where to stick the output files.
            `keep_headers`: Whether or not to print the headers in each output file.
        """
        print ('Creating Export paths')
        if not os.path.exists(output_path):
            os.makedirs(output_path)
        if not os.path.exists('Export'):
            os.makedirs('Export')

        reader = csv.reader(open(filehandler,'r'), delimiter=delimiter)
        current_piece = 1
        current_out_path = os.path.join(
             output_path,
             output_name_template  % current_piece
        )

        current_out_writer = csv.writer(open(current_out_path, 'w'), delimiter=delimiter)
        current_limit = row_limit
        if keep_headers:
            headers = next(reader)
            current_out_writer.writerow(headers)
        for i, row in enumerate(reader):
            if i + 1 > current_limit:
                current_piece += 1
                current_limit = row_limit * current_piece
                current_out_path = os.path.join(
                   output_path,
                   output_name_template  % current_piece
                )
                current_out_writer = csv.writer(open(current_out_path, 'w'), delimiter=delimiter)
                if keep_headers:
                    current_out_writer.writerow(headers)
            current_out_writer.writerow(row)
        return filehandler


    # join multiple csv to a single workbook with diff worksheets
    def split_join():
        print ('Joining files')
        # Returns the same day of last month if possible otherwise end of month
        # (eg: March 31st->29th Feb an July 31st->June 30th)
        last_month = datetime.now() - relativedelta(months=1)
        # Create string of month name and year...
        text = format(last_month, '%B %Y')
        prev_mon = '['+text+']'

        #print(prev_mon)
        fname = './Export/'+'Payments'+prev_mon+'.xlsx'
        writer = pd.ExcelWriter(fname, engine='xlsxwriter')
        folders = next(os.walk('.'))[1]
        for host in folders:
            Path = os.path.join(os.getcwd(), host)
            for f in glob.glob(os.path.join(Path, "Payments_*.csv")):
                #print(f)
                df = pd.read_csv(f,sep="\t",low_memory=False)
                df.to_excel(writer, index=False,sheet_name=os.path.basename(f)[:31])
            writer.save()
        return fname



    def styl():
        import openpyxl
        from openpyxl.reader.excel import load_workbook
        from openpyxl.styles import Font, Alignment 
        from openpyxl.utils import get_column_letter
        print('Applying')
        fname = split_join()
        print(fname)
        wb = load_workbook(fname)
        for ws in wb.worksheets:
            mr = ws.max_row
            mc = ws.max_column
            for cell in ws[mr:mc]:
                cell.font = Font(size=11)
            for cell in ws["1:1"]:
                cell.font = Font(size=12)
                cell.style = 'Accent1'
                cell.alignment = Alignment(wrapText='True',horizontal='center')
            for col in range(1, 30):
                ws.column_dimensions[(get_column_letter(col))].width = 15
                ws.freeze_panes = "A2"
        wb.save(fname)
    cleanup()
    split()
    split_join()
    styl()
    cleanup()

def chg_and_adj():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing Charges and Adjustments Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_chg_and_adj
    fs_ext = ".csv"


    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)


    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'

    #print(prev_mon)
    fd_name = './Export/'+'Charges and Adjustments'+prev_mon+'.xlsx'

    #main prog
    with alive_bar(total=100, manual='True',title='Charges and Adjustments',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def cct():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing CCT Info Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_cct
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)



    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'

    #print(prev_mon)
    fd_name = './Export/'+'CCT Information'+prev_mon+'.xlsx'

    #main prog
    with alive_bar(total=100, manual=True,title='CCT Information',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        # Columns: property_code date  serial_no   product applicationVersion  manufacturer    mfg_serial_no
        ws['A1'] = 'Property Code'
        ws['B1'] = 'Date'
        ws['C1'] = 'Serial No'
        ws['D1'] = 'Product'
        ws['E1'] = 'Application Version'
        ws['F1'] = 'Manufacturer'
        ws['G1'] = 'MFG Serial No'

        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def discp_rates():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing Discrepant Rates')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_discp_rates
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)



    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'

    #print(prev_mon)
    fd_name = './Export/'+'Discrepant Rates'+prev_mon+'.xlsx'

    #main prog
    with alive_bar(total=100, manual=True,title='Discrepant Rates',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def dnr1():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing DNR 1 Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_dnr1
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)



    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'

    #print(prev_mon)
    fd_name = './Export/'+'DNR 1'+prev_mon+'.xlsx'
    #main prog
    with alive_bar(total=100, manual=True,title='DNR 1',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def dnr2():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing DNR 2 Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_dnr2
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)



    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'

    #print(prev_mon)
    fd_name = './Export/'+'DNR 2'+prev_mon+'.xlsx'

    #main prog
    with alive_bar(total=100, manual=True,title='DNR 1',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"


        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def dnr3():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing DNR 3 Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_dnr3
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)

    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'

    #print(prev_mon)
    fd_name = './Export/'+'DNR 3'+prev_mon+'.xlsx'
    #main prog
    with alive_bar(total=100, manual=True,title='DNR 1',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"
        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def grts_and_gst():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing GRATIS & GSTCERT Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_grts_and_gst
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)

    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'

    #print(prev_mon)
    fd_name = './Export/'+'GRATIS and GSTCERT'+prev_mon+'.xlsx'

    #main prog
    with alive_bar(total=100, manual=True,title='GRATIS and GSTCERT 1',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def gst_email():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing Guest Email Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_gst_email
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)



    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'
    #print(prev_mon)
    fd_name = './Export/'+'Guest Email'+prev_mon+'.xlsx'

    #main prog
    with alive_bar(total=100, manual=True,title='Guest Email',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def lldb():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing LLDB Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_lldb
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)


    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'
    #print(prev_mon)
    fd_name = './Export/'+'LLDB'+prev_mon+'.xlsx'

    #main prog
    with alive_bar(total=100, manual=True,title='LLDB',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep='\t',header=0,names=['Property Code','Confirmation Number', 'Checkin Date','Checkout Date','Company Name','Total Charges','External Confirmation No'])
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def pay_and_ref():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing Payments & Refunds Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_pay_and_ref
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)

    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'
    #print(prev_mon)
    fd_name = './Export/'+'Payments and Refunds'+prev_mon+'.xlsx'

    #main prog
    with alive_bar(total=100, manual=True,title='Payments and Refunds',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def prop_over():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing Property Overview Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_prop_over
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)


    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'
    #print(prev_mon)
    fd_name = './Export/'+'Property Overview'+prev_mon+'.xlsx'
    #main prog
    with alive_bar(total=100, manual=True,title='Property Overview',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def room_moves():
    import openpyxl
    from openpyxl.reader.excel import load_workbook
    from openpyxl.styles import Font, Alignment 
    from openpyxl.utils import get_column_letter
    print('Processing Room Moves Report')
    ## DECLARE VARIABLES
    #Source vars
    fs_path = os.getcwd()
    fs_name = v_room_moves
    fs_ext = ".csv"

    ##Create working directory
    wd_name = "Export"
    wd = os.getcwd()
    #exp_dir=os.mkdir(wd_name)
    if not os.path.exists(wd_name):
        os.makedirs(wd_name)


    #Dest vars
    #fd_path = fs_path
    #fd_name = wd_name
    #fd_ext = ".xlsx"
    # Returns the same day of last month if possible otherwise end of month
    # (eg: March 31st->29th Feb an July 31st->June 30th)
    last_month = datetime.now() - relativedelta(months=1)
    # Create string of month name and year...
    text = format(last_month, '%B %Y')
    prev_mon = '['+text+']'
    #print(prev_mon)
    fd_name = './Export/'+'Room Moves'+prev_mon+'.xlsx'

    #main prog
    with alive_bar(total=100, manual=True,title='Room Moves',theme='smooth',bar='blocks',spinner='classic') as bar:   # default setting

        ## Report 1 - 
        #convert csv to xlsx using pandas lib
        read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext,sep="\t")
        bar(.10)
        read_file.to_excel (''r''+fd_name, index = None, header=True)
        bar(.20)
        #read xlsx
        #assign the excel file to wb() variable
        wb=load_workbook(fd_name)
        bar(.30)

        #assign the worksheet of the workbook to a ws() variable
        ws=wb.active
        bar(.40)
        mr = ws.max_row
        mc = ws.max_column
        for cell in ws[mr:mc]:
            cell.font = Font(size=11)
        bar(.60)
        # Set header row style
        for cell in ws["1:1"]:
            cell.font = Font(size=12)
            cell.style = 'Accent1'
            cell.alignment = Alignment(wrapText='True',horizontal='center')
        bar(.80)
        #set column width to 15 with loop
        for col in range(1, 30):
            ws.column_dimensions[(get_column_letter(col))].width = 15
            ws.freeze_panes = "A2"

        #Save the excel file
        bar(1)
        wb.save(fd_name)
        
    print('Export Completed\n')

def clean():
    if os.path.exists('Export'):
        shutil.rmtree('Export')

def all():
    print('Processing all reports')
    clean()
    #payments()
    cct()
    chg_and_adj()
    discp_rates()
    dnr1()
    dnr2()
    dnr3()
    grts_and_gst()
    gst_email()
    lldb()
    pay_and_ref()
    prop_over()
    room_moves()
    


dispatcher = {
    '1': payments,
    '2': cct,
    '3': chg_and_adj,
    '4': discp_rates,
    '5': dnr1,
    '6': dnr2,
    '7': dnr3,
    '8': grts_and_gst,
    '9': gst_email,
    '10': lldb,
    '11': pay_and_ref,
    '12': prop_over,
    '13': room_moves,
    'all': all
}
print('Following choices are available:\n\
    1 - Payments\n\
    2 - CCT\n\
    3 - Charges and adjustments\n\
    4 - Discrepant rates\n\
    5 - DNR 1\n\
    6 - DNR 2\n\
    7 - DNR 3\n\
    8 - GRTS and GSTCRT\n\
    9 - Guest Email\n\
    10 - LLDB\n\
    11 - Payments and Refunds\n\
    12 - Prop overview\n\
    13 - Room Moves\n\
    ')

action = input('Option: - ')

dispatcher[action]()

