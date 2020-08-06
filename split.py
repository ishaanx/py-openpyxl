import os
import xlsxwriter
import glob
import csv
import shutil
import pandas as pd
from datetime import datetime
from dateutil.relativedelta import relativedelta
from time import sleep
import string
from alive_progress import alive_bar
import sys
import subprocess
import time



def wrapper():
    output_path='./temp'

    def cleanup():
        print('Cleaning Up temp files')
        if os.path.exists(output_path):
            shutil.rmtree(output_path)


    def split(filehandler, delimiter=',', row_limit=300000, 
        output_name_template='Payments_%s.csv',  keep_headers=True):
        """
        Splits a CSV file into multiple pieces.
        Arguments:
            `row_limit`: The number of rows you want in each output file. 10,000 by default.
            `output_name_template`: A %s-style template for the numbered output files.
            `output_path`: Where to stick the output files.
            `keep_headers`: Whether or not to print the headers in each output file.
        """
        print ('Creating Export paths')
        if not os.path.exists(output_path):
            os.makedirs(output_path)
        if not os.path.exists('Export'):
            os.makedirs('Export')

        reader = csv.reader(filehandler, delimiter=delimiter)
        current_piece = 1
        current_out_path = os.path.join(
             output_path,
             output_name_template  % current_piece
        )

        current_out_writer = csv.writer(open(current_out_path, 'w'), delimiter=delimiter)
        current_limit = row_limit
        if keep_headers:
            headers = next(reader)
            current_out_writer.writerow(headers)
        for i, row in enumerate(reader):
            if i + 1 > current_limit:
                current_piece += 1
                current_limit = row_limit * current_piece
                current_out_path = os.path.join(
                   output_path,
                   output_name_template  % current_piece
                )
                current_out_writer = csv.writer(open(current_out_path, 'w'), delimiter=delimiter)
                if keep_headers:
                    current_out_writer.writerow(headers)
            current_out_writer.writerow(row)



    # join multiple csv to a single workbook with diff worksheets
    def split_join():
        print ('Joining files')
        # Returns the same day of last month if possible otherwise end of month
        # (eg: March 31st->29th Feb an July 31st->June 30th)
        last_month = datetime.now() - relativedelta(months=1)
        # Create string of month name and year...
        text = format(last_month, '%B %Y')
        prev_mon = '['+text+']'

        #print(prev_mon)
        fname = './Export/'+'Payments'+prev_mon+'.xlsx'
        writer = pd.ExcelWriter(fname, engine='xlsxwriter')
        folders = next(os.walk('.'))[1]
        for host in folders:
            Path = os.path.join(os.getcwd(), host)
            for f in glob.glob(os.path.join(Path, "Payments_*.csv")):
                #print(f)
                df = pd.read_csv(f,sep="\t",low_memory=False)
                df.to_excel(writer, index=False,sheet_name=os.path.basename(f)[:31])
            writer.save()
        return fname



    def styl():
        import openpyxl
        from openpyxl.reader.excel import load_workbook
        from openpyxl.styles import Font, Alignment 
        from openpyxl.utils import get_column_letter
        print('Applying')
        fname = split_join()
        print(fname)
        wb = load_workbook(fname)
        for ws in wb.worksheets:
            mr = ws.max_row
            mc = ws.max_column
            for cell in ws[mr:mc]:
                cell.font = Font(size=11)
            for cell in ws["1:1"]:
                cell.font = Font(size=12)
                cell.style = 'Accent1'
                cell.alignment = Alignment(wrapText='True',horizontal='center')
            for col in range(1, 54):
                ws.column_dimensions[(get_column_letter(col))].width = 15
                ws.freeze_panes = "A2"
        wb.save(fname)


    cleanup()
    split(open('Payments-Jun.csv', 'r'))
    split_join()
    styl()
    cleanup()


wrapper()
