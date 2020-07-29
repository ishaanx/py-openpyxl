## Script for excel manipulation. 
## First convert the input csv to excel
## Open the xlsx file
## Set Fonts, Styles, etc.
## Correct the headers
## Save the report

## Libraries used:
## < alive-progress > To show progress bar
## < Pandas > TO convert csv/tsv to xlsx
## < Openpyxl > For xlsx manipulation
## < os > Provides basic functions

import os
import pandas as pd
from openpyxl import *
from openpyxl.styles import *
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from time import sleep
import string
from alive_progress import alive_bar

print('Processing Gratis & GSTCERT Report')
## DECLARE VARIABLES
#Source vars
fs_path = os.getcwd()
fs_name = "test"
fs_ext = ".csv"

##Create working directory
wd_name = "Export"
wd = os.getcwd()
#exp_dir=os.mkdir(wd_name)
if not os.path.exists(wd_name):
    os.makedirs(wd_name)


#Dest vars
fd_path = fs_path
fd_name = wd_name
fd_ext = ".xlsx"

#main prog
with alive_bar(5, manual=True,title='Report 1',theme='smooth',bar='blocks') as bar:   # default setting

	## Report 1 - 
	#convert csv to xlsx using pandas lib
	read_file = pd.read_csv (''r''+fs_path+'/'+fs_name+fs_ext)
	bar(.10)
	read_file.to_excel (''r''+fd_path+'/'+fd_name+'/'+fs_name+fd_ext, index = None, header=True)
	bar(.20)
	#read xlsx
	#assign the excel file to wb() variable
	wb=load_workbook(fd_path+'/'+fd_name+'/'+fs_name+fd_ext)
	bar(.30)

	#assign the worksheet of the workbook to a ws() variable
	ws=wb.active
	bar(.40)
	mr = ws.max_row
	mc = ws.max_column
	for cell in ws[mr:mc]:
		cell.font = Font(size=11)
	bar(.60)
	# Set header row style
	for cell in ws["1:1"]:
		cell.font = Font(size=12)
		cell.style = 'Accent1'
	bar(.80)
	#set column width to 15 with loop
	for col in range(1, 54):
		ws.column_dimensions[(get_column_letter(col))].width = 15


	#Save the excel file
	bar(1)
	wb.save(fd_path+'/'+fd_name+'/'+fs_name+fd_ext)
	bar()
print('Export Completed')
